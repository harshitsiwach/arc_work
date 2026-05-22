import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_pitch_deck_spreadsheet():
    # Initialize workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    font_family = "Segoe UI"
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Slate
    subheader_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid") # Medium Gray
    accent_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid") # Soft Light Gray
    success_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Soft light green
    
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    italic_font = Font(name=font_family, size=9, italic=True, color="555555")
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    thin_side = Side(border_style="thin", color="D1D5DB")
    double_side = Side(border_style="double", color="374151")
    
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_total = Border(top=thin_side, bottom=double_side)

    # ────────────────────────────────────────────────────────
    # TAB 1: SLIDE OUTLINE
    # ────────────────────────────────────────────────────────
    ws1 = wb.create_sheet(title="1. Slide Outline")
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1.merge_cells("A1:D1")
    title_cell = ws1["A1"]
    title_cell.value = "Arc Work Pitch Deck - Slide Contents"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = align_center
    ws1.row_dimensions[1].height = 40
    
    # Headers
    headers_ws1 = ["Slide No.", "Slide Title", "Core Message / Subtitle", "Key Bullet Points / Visual Notes"]
    for col_idx, header in enumerate(headers_ws1, 1):
        cell = ws1.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = subheader_fill
        cell.alignment = align_center if col_idx == 1 else align_left
        cell.border = border_all
    ws1.row_dimensions[2].height = 25
    
    # Slide data
    slides_data = [
        (
            "Slide 1", 
            "Executive Summary", 
            "The On-chain Operating System for Internet Creators and AI Workers",
            "• Built on Arc L2 Blockchain optimized for consumer Web3 apps.\n"
            "• Employs Circle USDC/EURC developer-controlled wallets for seedless login.\n"
            "• Platform cuts set to ultra-competitive 2.5% platform fee.\n"
            "• Metric traction: 2,840+ creators registered, 14.2K+ completed orders."
        ),
        (
            "Slide 2", 
            "The Problem", 
            "Centralized bottlenecks restrict freelance gig and creator economic growth",
            "• Rent Extraction: Upwork, Fiverr, Whop charge 10% to 30% fees plus high FX fees.\n"
            "• Payment Hold: Payout holds typically take 7-14 days to settle to bank accounts.\n"
            "• Lock-In: Reputations are locked inside single platform silos.\n"
            "• AI workers: AI agents lack wallet rails to accept payouts and execute contracts."
        ),
        (
            "Slide 3", 
            "The Solution", 
            "Low-Fee, Instant onchain settlements backed by AI verification escrows",
            "• Low fee & instant payout (<1 second finality via Circle USDC on Arc L2).\n"
            "• Smart Escrows: EIP-712 templates hold funds securely until delivery.\n"
            "• AI Validation: OpenAI Vision API automatically validates deliverables against terms.\n"
            "• Open Portability: Portable ERC-8004 identity and on-chain ratings registry."
        ),
        (
            "Slide 4", 
            "Key Product Verticals", 
            "A unified suite of decentralized creative tools",
            "• Explore Portal (/explore): Selling and buying digital products and tools.\n"
            "• Freelance Gigs (/dashboard/marketplace): Escrow locked freelance board.\n"
            "• AI Agents Hub (/agents): Wizard to build, configure, and lease bots.\n"
            "• Gated Courses (/dashboard/courses): Video lessons locked behind x402 micropayments.\n"
            "• Bridge & Swap (/dashboard/bridge): Circle CCTP cross-chain bridge widget."
        ),
        (
            "Slide 5", 
            "Technical Architecture", 
            "Seamless integrations of Web3 wallets, smart contracts, and AI",
            "• Circle Developer-Controlled Wallets (DCWs): Google/Email OAuth seedless login.\n"
            "• Circle Smart Contract Platform (SCP): Programmable EIP-712 escrow deployment.\n"
            "• AI Validation Pipeline: Node backend + OpenAI Vision API checks deliverables.\n"
            "• x402 Micropayment Gating: Middleware validating transaction headers on-chain."
        ),
        (
            "Slide 6", 
            "Market Opportunity", 
            "High creator fee savings and exponential growth alignment",
            "• Target markets: Creator economy ($250B+ valuation) and Gig work ($10B+).\n"
            "• Competitive Fee Savings: Comparison matrix pitting 2.5% fee against legacy networks.\n"
            "• Visualizing the fee savings cumulative earnings simulator graph."
        ),
        (
            "Slide 7", 
            "Business Model", 
            "Platform monetization with minimal transaction friction",
            "• 2.5% Payout Fee: Split as 1.5% platform net treasury and 1.0% ecosystem gas relayer.\n"
            "• AI Execution Surcharge: Minor $0.01 - $0.05 surcharge per agent runtime step.\n"
            "• On-chain Subscriptions: Monthly SaaS membership tiers via ERC-8191 standard."
        ),
        (
            "Slide 8", 
            "Future Roadmap", 
            "Direct path to Mainnet launch and fiat integration layers",
            "• Phase 1 (Completed): Core escrow contracts, Circle wallets, OpenAI Vision verification.\n"
            "• Phase 2 (Completed): Navigation IA overhaul, sidebars, settings, and x402 learning player.\n"
            "• Phase 3 (Planned): Mainnet launch, Circle fiat on/off-ramp widgets, cross-platform ZK reputation."
        ),
    ]
    
    current_row = 3
    for s_num, s_title, s_msg, s_bullets in slides_data:
        ws1.cell(row=current_row, column=1, value=s_num).alignment = align_center
        ws1.cell(row=current_row, column=1).font = bold_font
        ws1.cell(row=current_row, column=1).border = border_all
        
        ws1.cell(row=current_row, column=2, value=s_title).alignment = align_left
        ws1.cell(row=current_row, column=2).font = bold_font
        ws1.cell(row=current_row, column=2).border = border_all
        
        ws1.cell(row=current_row, column=3, value=s_msg).alignment = align_left
        ws1.cell(row=current_row, column=3).font = regular_font
        ws1.cell(row=current_row, column=3).border = border_all
        
        cell_bullets = ws1.cell(row=current_row, column=4, value=s_bullets)
        cell_bullets.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_bullets.font = regular_font
        cell_bullets.border = border_all
        
        ws1.row_dimensions[current_row].height = 90
        current_row += 1

    # Adjust columns widths for Tab 1
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 45
    ws1.column_dimensions["D"].width = 75

    # ────────────────────────────────────────────────────────
    # TAB 2: FINANCIAL MODEL
    # ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="2. Financial Model")
    ws2.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws2.merge_cells("A1:C1")
    title_cell = ws2["A1"]
    title_cell.value = "Arc Work Treasury Projections Calculator"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = align_center
    ws2.row_dimensions[1].height = 40
    
    # Section Header: Assumptions (Editable)
    ws2.merge_cells("A2:C2")
    sub_cell = ws2["A2"]
    sub_cell.value = "INPUT ASSUMPTIONS (Edit these cells to change projections)"
    sub_cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    sub_cell.fill = subheader_fill
    sub_cell.alignment = align_left
    ws2.row_dimensions[2].height = 25
    
    # Write Assumptions Rows
    assumptions = [
        ("Monthly Active Users (MAUs)", 10000, "#,##0"),
        ("Avg Transaction Size in USDC", 150, "$#,##0.00"),
        ("Avg Transactions / User / Month", 1.0, "0.0"),
        ("Platform Transaction Fee Rate", 0.025, "0.0%"),
        ("Treasury Share of Transaction Fee", 0.60, "0.0%"),  # 1.5% out of 2.5% is 60%
        ("AI Agent Runs / User / Month", 5, "#,##0"),
        ("Avg AI Agent Run Fee in USDC", 0.02, "$#,##0.00"),
        ("Premium SaaS Subscription Conversion Rate", 0.05, "0.0%"),
        ("Premium SaaS Monthly Subscription Fee", 20, "$#,##0.00")
    ]
    
    for i, (label, val, num_format) in enumerate(assumptions, 3):
        ws2.cell(row=i, column=1, value=label).font = bold_font
        ws2.cell(row=i, column=1).alignment = align_left
        ws2.cell(row=i, column=1).border = border_all
        
        val_cell = ws2.cell(row=i, column=2, value=val)
        val_cell.font = bold_font
        val_cell.alignment = align_right
        val_cell.number_format = num_format
        val_cell.fill = success_fill # highlight editable inputs
        val_cell.border = border_all
        
        ws2.cell(row=i, column=3, value="← EDITABLE VARIABLE").font = italic_font
        ws2.cell(row=i, column=3).alignment = align_left
        ws2.cell(row=i, column=3).border = border_all
        
        ws2.row_dimensions[i].height = 22

    # Section Header: Projections (Calculated)
    ws2.merge_cells("A13:C13")
    proj_header_cell = ws2["A13"]
    proj_header_cell.value = "MONTHLY PLATFORM TREASURY PROJECTIONS"
    proj_header_cell.font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    proj_header_cell.fill = subheader_fill
    proj_header_cell.alignment = align_left
    ws2.row_dimensions[13].height = 25
    
    # Write Projections Rows (with formulas referencing assumptions)
    projections = [
        ("Total Monthly Transaction Volume", "=B3*B4*B5", "$#,##0.00"),
        ("Gross Transaction Fees Generated", "=B14*B6", "$#,##0.00"),
        ("Net Transaction Payouts to Treasury", "=B15*B7", "$#,##0.00"),
        ("AI Agent Execution Payouts to Treasury", "=B3*B8*B9", "$#,##0.00"),
        ("ERC-8191 Subscription Payouts to Treasury", "=B3*B10*B11", "$#,##0.00"),
    ]
    
    for idx, (label, formula, num_format) in enumerate(projections, 14):
        ws2.cell(row=idx, column=1, value=label).font = regular_font
        ws2.cell(row=idx, column=1).alignment = align_left
        ws2.cell(row=idx, column=1).border = border_all
        
        cell = ws2.cell(row=idx, column=2, value=formula)
        cell.font = bold_font
        cell.alignment = align_right
        cell.number_format = num_format
        cell.border = border_all
        
        ws2.cell(row=idx, column=3, value="").border = border_all # empty comment column
        ws2.row_dimensions[idx].height = 22
        
    # Total Platform Revenue Row (Row 19)
    total_row = 19
    ws2.cell(row=total_row, column=1, value="TOTAL MONTHLY PLATFORM REVENUE").font = bold_font
    ws2.cell(row=total_row, column=1).alignment = align_left
    ws2.cell(row=total_row, column=1).fill = accent_fill
    ws2.cell(row=total_row, column=1).border = border_total
    
    total_cell = ws2.cell(row=total_row, column=2, value="=B16+B17+B18")
    total_cell.font = Font(name=font_family, size=11, bold=True, color="10B981") # Accent green text
    total_cell.alignment = align_right
    total_cell.number_format = "$#,##0.00"
    total_cell.fill = accent_fill
    total_cell.border = border_total
    
    ws2.cell(row=total_row, column=3, value="Net monthly treasury USDC").font = italic_font
    ws2.cell(row=total_row, column=3).alignment = align_left
    ws2.cell(row=total_row, column=3).fill = accent_fill
    ws2.cell(row=total_row, column=3).border = border_total
    ws2.row_dimensions[total_row].height = 26

    # Adjust column widths for Tab 2
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 25

    # ────────────────────────────────────────────────────────
    # TAB 3: FEE SAVINGS
    # ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet(title="3. Creator Fee Savings")
    ws3.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws3.merge_cells("A1:I1")
    title_cell = ws3["A1"]
    title_cell.value = "Creator Net Earnings and Fee Savings Comparison Calculator"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = align_center
    ws3.row_dimensions[1].height = 40
    
    # Table Headers
    headers_ws3 = [
        "Tx Volume", 
        "Upwork Fee (20%)", "Upwork Net", 
        "Whop Fee (5%)", "Whop Net", 
        "Arc Work Fee (2.5%)", "Arc Work Net", 
        "Arc Savings vs Upwork", "Arc Savings vs Whop"
    ]
    for col_idx, header in enumerate(headers_ws3, 1):
        cell = ws3.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = subheader_fill
        cell.alignment = align_center
        cell.border = border_all
    ws3.row_dimensions[2].height = 25
    
    # Volumes list for calculator
    volumes = [1000, 2500, 5000, 7500, 10000, 15000, 20000, 25000, 30000, 40000, 50000]
    
    for row_idx, vol in enumerate(volumes, 3):
        # Column A: Volume
        cell_vol = ws3.cell(row=row_idx, column=1, value=vol)
        cell_vol.font = bold_font
        cell_vol.alignment = align_right
        cell_vol.number_format = "$#,##0"
        cell_vol.border = border_all
        
        # Column B: Upwork Fee
        cell_up_fee = ws3.cell(row=row_idx, column=2, value=f"=A{row_idx}*0.20")
        cell_up_fee.font = regular_font
        cell_up_fee.alignment = align_right
        cell_up_fee.number_format = "$#,##0.00"
        cell_up_fee.border = border_all
        
        # Column C: Upwork Net
        cell_up_net = ws3.cell(row=row_idx, column=3, value=f"=A{row_idx}-B{row_idx}")
        cell_up_net.font = regular_font
        cell_up_net.alignment = align_right
        cell_up_net.number_format = "$#,##0.00"
        cell_up_net.border = border_all
        
        # Column D: Whop Fee
        cell_whop_fee = ws3.cell(row=row_idx, column=4, value=f"=A{row_idx}*0.05")
        cell_whop_fee.font = regular_font
        cell_whop_fee.alignment = align_right
        cell_whop_fee.number_format = "$#,##0.00"
        cell_whop_fee.border = border_all
        
        # Column E: Whop Net
        cell_whop_net = ws3.cell(row=row_idx, column=5, value=f"=A{row_idx}-D{row_idx}")
        cell_whop_net.font = regular_font
        cell_whop_net.alignment = align_right
        cell_whop_net.number_format = "$#,##0.00"
        cell_whop_net.border = border_all
        
        # Column F: Arc Work Fee
        cell_arc_fee = ws3.cell(row=row_idx, column=6, value=f"=A{row_idx}*0.025")
        cell_arc_fee.font = bold_font
        cell_arc_fee.alignment = align_right
        cell_arc_fee.number_format = "$#,##0.00"
        cell_arc_fee.border = border_all
        
        # Column G: Arc Work Net (Highlight in light green)
        cell_arc_net = ws3.cell(row=row_idx, column=7, value=f"=A{row_idx}-F{row_idx}")
        cell_arc_net.font = bold_font
        cell_arc_net.alignment = align_right
        cell_arc_net.number_format = "$#,##0.00"
        cell_arc_net.fill = success_fill
        cell_arc_net.border = border_all
        
        # Column H: Arc Savings vs Upwork
        cell_sav_up = ws3.cell(row=row_idx, column=8, value=f"=B{row_idx}-F{row_idx}")
        cell_sav_up.font = Font(name=font_family, size=11, bold=True, color="10B981")
        cell_sav_up.alignment = align_right
        cell_sav_up.number_format = "$#,##0.00"
        cell_sav_up.border = border_all
        
        # Column I: Arc Savings vs Whop
        cell_sav_whop = ws3.cell(row=row_idx, column=9, value=f"=D{row_idx}-F{row_idx}")
        cell_sav_whop.font = Font(name=font_family, size=11, bold=True, color="10B981")
        cell_sav_whop.alignment = align_right
        cell_sav_whop.number_format = "$#,##0.00"
        cell_sav_whop.border = border_all
        
        ws3.row_dimensions[row_idx].height = 20

    # Adjust columns widths for Tab 3
    for col in range(1, 10):
        col_letter = get_column_letter(col)
        ws3.column_dimensions[col_letter].width = 22

    # Save to file
    filepath = "PITCH_DECK_MODEL.xlsx"
    wb.save(filepath)
    print(f"Spreadsheet generated successfully and saved to: {filepath}")

if __name__ == "__main__":
    create_pitch_deck_spreadsheet()
